from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.application.ports.reporting import ShiftReportRendererPort


class OpenpyxlShiftReportRenderer(ShiftReportRendererPort):
    async def render(
        self,
        *,
        report_id,  # noqa: ARG002
        report_version: int,
        project_title: str,
        shift_title: str,
        shift_start_time,
        shift_end_time,
        actuality_status: str,
        generated_at,
        participants: tuple[dict[str, object], ...],
        owner_sections: tuple[dict[str, object], ...],
        external_owner_sections: tuple[dict[str, object], ...],
    ) -> bytes:
        workbook = Workbook()
        summary = workbook.active
        summary.title = "Сводка"
        self._write_summary(
            sheet=summary,
            project_title=project_title,
            shift_title=shift_title,
            shift_start_time=shift_start_time,
            shift_end_time=shift_end_time,
            report_version=report_version,
            actuality_status=actuality_status,
            generated_at=generated_at,
        )

        participants_sheet = workbook.create_sheet("Участники")
        self._write_participants(participants_sheet, participants)

        for section in owner_sections:
            owner_name = str(section["owner_display_name"])
            sheet = workbook.create_sheet(_sheet_name(owner_name))
            self._write_owner_resources(
                sheet=sheet,
                owner_display_name=owner_name,
                project_role=str(section.get("project_role", "")),
                shift_role=str(section.get("shift_role", "")),
                resources=tuple(section["resources"]),
            )

        if external_owner_sections:
            external_sheet = workbook.create_sheet("Внешние владельцы")
            self._write_external_owners(external_sheet, external_owner_sections)

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _write_summary(
        self,
        *,
        sheet,
        project_title: str,
        shift_title: str,
        shift_start_time,
        shift_end_time,
        report_version: int,
        actuality_status: str,
        generated_at,
    ) -> None:
        rows = [
            ("Проект", project_title),
            ("Смена", shift_title),
            ("Интервал смены", f"{shift_start_time} - {shift_end_time}"),
            ("Версия отчета", report_version),
            ("Сформирован", str(generated_at)),
            ("Актуальность", _translate_actuality_status(actuality_status)),
        ]
        sheet["A1"] = "Отчет по смене"
        _style_title(sheet["A1"])
        for index, (label, value) in enumerate(rows, start=3):
            sheet[f"A{index}"] = label
            sheet[f"B{index}"] = value
            _style_header(sheet[f"A{index}"])
        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 60

    def _write_participants(self, sheet, participants: tuple[dict[str, object], ...]) -> None:
        headers = [
            "Участник",
            "Телефон",
            "Почта",
            "Роль в проекте",
            "Роль на смене",
            "Время с",
            "Время по",
            "Прибыл",
            "Время прибытия",
            "Примечание",
        ]
        self._write_headers(sheet, headers)
        for row_index, participant in enumerate(participants, start=2):
            sheet.cell(row=row_index, column=1, value=_translate_display_text(participant["username"]))
            sheet.cell(row=row_index, column=2, value=_translate_display_text(participant["phone"]))
            sheet.cell(row=row_index, column=3, value=_translate_display_text(participant["email"]))
            sheet.cell(
                row=row_index,
                column=4,
                value=_translate_project_role(participant["project_role"]),
            )
            sheet.cell(
                row=row_index,
                column=5,
                value=_translate_project_role(participant["shift_role"]),
            )
            sheet.cell(row=row_index, column=6, value=str(participant["time_from"]))
            sheet.cell(row=row_index, column=7, value=str(participant["time_to"]))
        _set_column_widths(sheet, [28, 18, 28, 16, 16, 24, 24, 12, 16, 28])

    def _write_owner_resources(
        self,
        *,
        sheet,
        owner_display_name: str,
        project_role: str,
        shift_role: str,
        resources: tuple[dict[str, object], ...],
    ) -> None:
        sheet["A1"] = f"Ресурсы: {_translate_display_text(owner_display_name)}"
        _style_title(sheet["A1"])
        if project_role:
            sheet["A2"] = f"Роль в проекте: {_translate_project_role(project_role)}"
        if shift_role:
            sheet["B2"] = f"Роль на смене: {_translate_project_role(shift_role)}"
        headers = [
            "Название",
            "Тип",
            "Описание",
            "Размер",
            "Владелец",
            "Время с",
            "Время по",
            "Оборудование доставлено",
            "Проверил",
            "Примечание",
        ]
        self._write_headers(sheet, headers, row=4)
        for row_index, resource in enumerate(resources, start=5):
            sheet.cell(row=row_index, column=1, value=_translate_display_text(resource["title"]))
            sheet.cell(row=row_index, column=2, value=_translate_display_text(resource["type"]))
            sheet.cell(
                row=row_index,
                column=3,
                value=_translate_display_text(resource["description"]),
            )
            sheet.cell(row=row_index, column=4, value=_translate_display_text(resource["size"]))
            sheet.cell(
                row=row_index,
                column=5,
                value=_translate_display_text(resource["owner_display_name"]),
            )
            sheet.cell(row=row_index, column=6, value=str(resource["time_from"]))
            sheet.cell(row=row_index, column=7, value=str(resource["time_to"]))
        _set_column_widths(sheet, [28, 18, 40, 12, 24, 24, 24, 14, 18, 28])

    def _write_external_owners(self, sheet, sections: tuple[dict[str, object], ...]) -> None:
        row = 1
        for section in sections:
            sheet.cell(
                row=row,
                column=1,
                value=f"Внешний владелец: {_translate_display_text(section['owner_display_name'])}",
            )
            _style_title(sheet.cell(row=row, column=1))
            row += 2
            headers = [
                "Название",
                "Тип",
                "Описание",
                "Размер",
                "Владелец",
                "Время с",
                "Время по",
                "Оборудование доставлено",
                "Проверил",
                "Примечание",
            ]
            self._write_headers(sheet, headers, row=row)
            row += 1
            for resource in section["resources"]:
                sheet.cell(row=row, column=1, value=_translate_display_text(resource["title"]))
                sheet.cell(row=row, column=2, value=_translate_display_text(resource["type"]))
                sheet.cell(
                    row=row,
                    column=3,
                    value=_translate_display_text(resource["description"]),
                )
                sheet.cell(row=row, column=4, value=_translate_display_text(resource["size"]))
                sheet.cell(
                    row=row,
                    column=5,
                    value=_translate_display_text(resource["owner_display_name"]),
                )
                sheet.cell(row=row, column=6, value=str(resource["time_from"]))
                sheet.cell(row=row, column=7, value=str(resource["time_to"]))
                row += 1
            row += 2
        _set_column_widths(sheet, [28, 18, 40, 12, 24, 24, 24, 14, 18, 28])

    def _write_headers(self, sheet, headers: Iterable[str], *, row: int = 1) -> None:
        for col_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row, column=col_index, value=header)
            _style_header(cell)


def _style_title(cell) -> None:
    cell.font = Font(bold=True, size=14)


def _style_header(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _set_column_widths(sheet, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[_column_letter(index)].width = width


def _column_letter(index: int) -> str:
    result = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _sheet_name(raw_name: str) -> str:
    cleaned = "".join("_" if ch in "\\/*?:[]" else ch for ch in raw_name).strip()
    if not cleaned:
        cleaned = "Участник"
    return cleaned[:31]


ROLE_LABELS = {
    "DIRECTOR": "Режиссер",
    "PROP_MASTER": "Реквизитор",
    "CAMERA": "Оператор",
    "SOUND": "Звукорежиссер",
    "LIGHT": "Осветитель",
    "ACTOR": "Актер",
    "UNKNOWN": "Не указано",
}

ACTUALITY_LABELS = {
    "ACTUAL": "Актуальный",
    "STALE": "Устаревший",
}

PLACEHOLDER_LABELS = {
    "Unknown user": "Неизвестный пользователь",
    "Missing phone": "Телефон не указан",
    "Missing email": "Почта не указана",
    "Unknown resource": "Неизвестный ресурс",
    "Missing description": "Описание не указано",
}


def _translate_project_role(value: object) -> object:
    if value is None:
        return None
    return ROLE_LABELS.get(str(value), value)


def _translate_actuality_status(value: object) -> object:
    if value is None:
        return None
    return ACTUALITY_LABELS.get(str(value), value)


def _translate_display_text(value: object) -> object:
    if value is None:
        return None
    return PLACEHOLDER_LABELS.get(str(value), value)
