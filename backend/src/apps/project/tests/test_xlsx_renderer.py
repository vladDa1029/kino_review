import asyncio
from datetime import UTC, datetime
from io import BytesIO
from uuid import uuid4

from openpyxl import load_workbook

from app.infrastructure.reporting.xlsx import OpenpyxlShiftReportRenderer


def test_shift_report_renderer_uses_russian_labels() -> None:
    renderer = OpenpyxlShiftReportRenderer()
    report_bytes = asyncio.run(
        renderer.render(
            report_id=uuid4(),
            report_version=3,
            project_title="Большое кино",
            shift_title="Ночная смена",
            shift_start_time=datetime(2026, 4, 22, 20, 0, tzinfo=UTC),
            shift_end_time=datetime(2026, 4, 23, 2, 0, tzinfo=UTC),
            actuality_status="ACTUAL",
            generated_at=datetime(2026, 4, 22, 18, 30, tzinfo=UTC),
            participants=(
                {
                    "username": "Иван Иванов",
                    "phone": "+79990001122",
                    "email": "ivan@example.com",
                    "project_role": "ACTOR",
                    "shift_role": "CAMERA",
                    "time_from": datetime(2026, 4, 22, 20, 0, tzinfo=UTC),
                    "time_to": datetime(2026, 4, 23, 2, 0, tzinfo=UTC),
                },
            ),
            owner_sections=(
                {
                    "owner_display_name": "Иван Иванов",
                    "project_role": "ACTOR",
                    "shift_role": "CAMERA",
                    "resources": (
                        {
                            "title": "Sony A7S III",
                            "type": "camera",
                            "description": "Основная камера",
                            "size": "M",
                            "owner_display_name": "Иван Иванов",
                            "time_from": datetime(2026, 4, 22, 20, 0, tzinfo=UTC),
                            "time_to": datetime(2026, 4, 23, 2, 0, tzinfo=UTC),
                        },
                    ),
                },
            ),
            external_owner_sections=(
                {
                    "owner_display_name": "Павел",
                    "resources": (
                        {
                            "title": "ARRI SkyPanel",
                            "type": "light",
                            "description": "Свет",
                            "size": "L",
                            "owner_display_name": "Павел",
                            "time_from": datetime(2026, 4, 22, 20, 0, tzinfo=UTC),
                            "time_to": datetime(2026, 4, 23, 2, 0, tzinfo=UTC),
                        },
                    ),
                },
            ),
        )
    )

    workbook = load_workbook(BytesIO(report_bytes))

    assert workbook.sheetnames == ["Сводка", "Участники", "Иван Иванов", "Внешние владельцы"]

    summary_sheet = workbook["Сводка"]
    assert summary_sheet["A1"].value == "Отчет по смене"
    assert summary_sheet["A3"].value == "Проект"
    assert summary_sheet["A8"].value == "Актуальность"
    assert summary_sheet["B8"].value == "Актуальный"
    # Время переводится из UTC в московское (UTC+3) и форматируется как ДД.ММ.ГГГГ ЧЧ:ММ.
    assert summary_sheet["B5"].value == "22.04.2026 23:00 - 23.04.2026 05:00"
    assert summary_sheet["B7"].value == "22.04.2026 21:30"

    participants_sheet = workbook["Участники"]
    assert participants_sheet["A1"].value == "Участник"
    assert participants_sheet["D2"].value == "Актер"
    assert participants_sheet["E2"].value == "Оператор"
    assert participants_sheet["F2"].value == "22.04.2026 23:00"
    assert participants_sheet["G2"].value == "23.04.2026 05:00"

    owner_sheet = workbook["Иван Иванов"]
    assert owner_sheet["A1"].value == "Ресурсы: Иван Иванов"
    assert owner_sheet["A2"].value == "Роль в проекте: Актер"
    assert owner_sheet["B2"].value == "Роль на смене: Оператор"

    external_sheet = workbook["Внешние владельцы"]
    assert external_sheet["A1"].value == "Внешний владелец: Павел"
    assert external_sheet["A3"].value == "Название"
